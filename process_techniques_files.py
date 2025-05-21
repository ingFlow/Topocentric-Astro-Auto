import pd_automate
import datetime 
import julian
import swisseph as swe
import secondary_automate
import pssr_swiss_auto as pssr_auto
import transit_swiss_auto as transit_auto
import sra_auto
import pandas as pd
import lunar_auto as lunar
import re
import os
import csv
from timezonefinder import TimezoneFinder
import pytz
from constants import calc_planets_pof_houses_labelled, PLANETS
from aspects_base import calculate_obliquity
import ast
import json
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from constants import parse_selection_file, DATA_INPUT_DIR, SELECTIONS_DIR, aTechniqueType, get_technique_name, PLANET_ABBREVIATIONS, ALL_ASPECTS

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

class TechniqueType:
    PRIMARY_DIRECT = 0
    SECONDARY_DIRECT = 1
    PSSR = 2
    TRANSIT = 3
    LUNAR = 4

grid_aspects =[]
date_technique = -1
aspect_type = -1

def resetvars():
    global grid_aspects, date_technique, aspect_type
    grid_aspects = []
    date_technique = -1
    aspect_type = -1

def generate_grid_angular_aspects(filename, start_time, end_time, increment_seconds, list_dt_events, geo_positions: list[3], type: pd_automate.AspectType, technique: TechniqueType):
    global grid_aspects, date_technique, aspect_type
    date_technique = technique
    aspect_type = type
    temp_list_event = ['Time']

    for i in range(0,len(list_dt_events)):
        temp_list_event.append(f"{i}: {list_dt_events[i][0].strftime('%Y-%m-%d')}")
    temp_list_event.append('Count')
    grid_aspects.append(temp_list_event)
    
    current_time = start_time
    increment = timedelta(seconds=increment_seconds)
    
    while current_time <= end_time:
        print(f"working on: {current_time}....")
        append_grid_acceptable_angles(list_dt_events, julian.to_jd(current_time),geo_positions)
        current_time += increment

    # Handle the case where the last increment might exceed the end time
    if current_time > end_time:
        append_grid_acceptable_angles(list_dt_events, julian.to_jd(current_time),geo_positions)
    
    directory = os.path.dirname(filename)
    os.makedirs(directory, exist_ok=True)
    
    with open(f"{filename}.txt", "w") as file:
        for time in grid_aspects:
            file.write(f"{str(time)}\n")
    
def append_grid_acceptable_angles(list_dt_events, jd_radix : julian, geopos_natal: list[3]):
    formatted_datetime = julian.from_jd(jd_radix).strftime("%Y-%m-%d %H:%M:%S")
    temp_list_event = [formatted_datetime] 
    count = 0
    
    rad_houses_info = swe.houses(jd_radix, geopos_natal[0], geopos_natal[1], b'T')
    rad_planets_equatorial = pd_automate.calc_rad_planets_equatorial(jd_radix)
    rad_planets_pof_houses_labelled = calc_planets_pof_houses_labelled(jd_radix, geopos_natal)
    e = calculate_obliquity(jd_radix)

    event_index = 0
    for dt_event, event_id, geopos in list_dt_events:
        if date_technique == TechniqueType.PRIMARY_DIRECT:
            pd_auto_obj  = pd_automate.PD_Automate(jd_radix, julian.to_jd(dt_event), geopos_natal, rad_planets_pof_houses_labelled, rad_planets_equatorial, rad_houses_info, e)
            str_rad_dir_aspects, str_rad_conv_aspects = pd_auto_obj.get_aspects_str()
            str_all_directed_aspects = str_rad_dir_aspects + str_rad_conv_aspects   
        elif date_technique == TechniqueType.SECONDARY_DIRECT:
            secondary_obj = secondary_automate.Secondary_Auto(jd_radix, julian.to_jd(dt_event), geopos_natal[0], geopos_natal[1], e, rad_houses_info[1][2], rad_planets_pof_houses_labelled)
            str_rad_n_prog_aspects, str_rad_n_reg_aspects = secondary_obj.get_str_aspects()
            str_all_directed_aspects = str_rad_n_prog_aspects + '\n' + str_rad_n_reg_aspects
        elif date_technique == TechniqueType.PSSR:
            pssr_obj = pssr_auto.PSSR_Auto(julian.from_jd(jd_radix), dt_event, rad_planets_pof_houses_labelled)
            str_rad_dir_aspects, str_rad_conv_aspects = pssr_obj.get_str_aspects()
            str_all_directed_aspects = str_rad_dir_aspects + str_rad_conv_aspects 
        elif date_technique == TechniqueType.TRANSIT:
            transit_obj = transit_auto.Transit_Auto(jd_radix, julian.to_jd(dt_event), geopos, rad_planets_pof_houses_labelled)
            str_rad_dir_aspects, str_rad_conv_aspects = transit_obj.get_str_aspects()
            str_all_directed_aspects = str_rad_dir_aspects + str_rad_conv_aspects 
        elif date_technique == TechniqueType.SRA:
            sra_auto_obj = sra_auto.SRA_Auto(julian.from_jd(jd_radix), dt_event, geopos_natal,rad_planets_pof_houses_labelled)
            str_rad_dir_aspects, str_rad_conv_aspects = sra_auto_obj.get_str_aspects()
            str_all_directed_aspects = str_rad_dir_aspects + str_rad_conv_aspects 
            str_all_directed_aspects = str_all_directed_aspects.replace(")(", ")\n(")
        
    
        ''' if date_technique == TechniqueType.PRIMARY_DIRECT:
            count, str_acceptable_aspects = pd_automate.count_pd_score_acceptable_aspects(event_id, str_all_directed_aspects, count)
        else:'''
        count, str_acceptable_aspects = pd_automate.count_event_acceptable_aspects(event_id, str_all_directed_aspects, count, aspect_type)

        if str_acceptable_aspects == '':
            temp_list_event.append(f"{str(event_index)}")
        else:
            temp_list_event.append(str_acceptable_aspects)
    
        event_index += 1

    temp_list_event.append(count)

    global grid_aspects
    grid_aspects.append(temp_list_event)    
    return

def categorize_aspect(first, second, aspect):
    conj_asp = ['conjunction', 'opposition']
    maj_asp = ['trine', 'sextile', 'square']
    ANGLES = ['H1', 'H4', 'H7', 'H10']

    if first in ANGLES and second in PLANETS :
        if aspect in conj_asp:
            return 'p_a_conj'
        elif aspect in maj_asp:
            return 'p_a_maj'
        else:
            return 'p_a_min'
    elif second in ANGLES and first in PLANETS:
        if aspect in conj_asp:
            return 'a_p_conj'
        elif aspect in maj_asp:
            return 'a_p_maj'
        else:
            return 'a_p_min'
    elif first.startswith('H') and second in PLANETS: #the other angle to house already evaluated so if its still H its not an angle
        if aspect in conj_asp:
            return 'p_h_conj'
        elif aspect in maj_asp:
            return 'p_h_maj'
        else:
            return 'p_h_min'
    elif second.startswith('H') and first in PLANETS: 
        if aspect in conj_asp:
            return 'h_p_conj'
        elif aspect in maj_asp:
            return 'h_p_maj'
        else:
            return 'h_p_min'
    elif first in PLANETS and second == 'Moon':
        if aspect in conj_asp:
            return 'mon_p_conj'
        elif aspect in maj_asp:
            return 'mon_p_maj'
        else:
            return 'mon_p_min'
    elif first in PLANETS and second in PLANETS:
        if aspect in conj_asp:
            return 'p_p_conj'
        elif aspect in maj_asp:
            return 'p_p_maj'
        else: 
            return 'p_p_min'
    else:
        return None

def count_extended_aspect_groups_txt(filename, technique):
    aspect_categories = {
        'p_a_conj': 0,
        'p_a_maj': 0,
        'p_a_allm': 0,
        'p_a_min': 0,
        'a_p_conj': 0,
        'a_p_maj': 0,
        'a_p_allm': 0,
        'a_p_min': 0,
        'p_h_conj': 0,
        'p_h_maj': 0,
        'p_h_allm': 0,
        'p_h_min': 0,
        'h_p_conj': 0,
        'h_p_maj': 0,
        'h_p_allm': 0,
        'h_p_min': 0,
        'mon_p_conj': 0,
        'mon_p_maj': 0,
        'mon_p_allm': 0,
        'mon_p_min': 0,
        'p_p_conj': 0,
        'p_p_maj': 0,
        'p_p_allm': 0,
        'p_p_all_m': 0,
        'e': 0,
    }
    results = []
    flag_1st_row = True

    with open(f"{filename}.txt", 'r') as infile:
        for row in infile:
            if flag_1st_row:
                flag_1st_row = False
            else:
                parts = eval(row.strip())
                time = parts[0]
                aspects = parts[1:-1]
                count_total = parts[-1]
                counts = {
                    'time': time
                }
                counts.update(aspect_categories.copy())  #Reset counts

                for aspect in aspects: 
                    if aspect.isdigit():
                        counts['e'] += 1
                    else:
                        aspects = aspect.split('\n')  # Split multiple aspects
                        for a in aspects:
                            parts = a.strip().strip('()').split(') (')
                            if len(parts) == 3:
                                first = parts[0].split(',')[0]  # Get the first planet info
                                second = parts[1].split(',')[0]  # Get the second planet info
                                asp_type = parts[2].split(',')[0] #Is it conjunction or quincunx etc.
                                category = categorize_aspect(first, second, asp_type)
                                if category:
                                    counts[category] += 1
                counts.update({'total_count':count_total})
                #{TODO decide whether to integrate p_p directions i put them here cause level_aspects is what will filter them out if need be}
                counts['p_a_allm'] = counts['p_a_conj'] + counts['p_a_maj']
                counts['a_p_allm'] = counts['a_p_conj'] + counts['a_p_maj']
                counts['p_h_allm'] = counts['p_h_conj'] + counts['p_h_maj']
                counts['h_p_allm'] = counts['h_p_conj'] + counts['h_p_maj']
                counts['mon_p_allm'] = counts['mon_p_conj'] + counts['mon_p_maj']
                counts['p_p_allm'] = counts['p_p_conj'] + counts['p_p_maj']
                counts['all_conj'] = counts['p_a_conj'] + counts['a_p_conj'] + counts['p_h_conj'] + counts['h_p_conj'] + counts['mon_p_conj'] + counts['p_p_conj']
                counts['p_a_mon_conj'] = counts['p_a_conj'] + counts['mon_p_conj']
                counts['a_p_p_a_conj'] = counts['a_p_conj'] + counts['p_a_conj']    
                counts['all_maj'] = counts['p_h_maj'] + counts['p_a_maj'] + counts['mon_p_maj'] + counts['a_p_maj'] + counts['h_p_maj'] + counts['p_p_maj']
                counts['all_m'] = counts['a_p_allm'] + counts['h_p_allm'] + counts['p_a_allm'] + counts['p_h_allm'] + counts['mon_p_allm'] + counts['p_p_allm']

                results.append(counts) 

    with open(f"{filename}COUNT.txt", 'w') as f:
        for index, count in enumerate(results):
            if technique == TechniqueType.PRIMARY_DIRECT:
                selected_categories = ['time', 'p_a_conj', 'p_a_maj', 'p_a_allm', 'p_a_min', 'a_p_conj', 'a_p_maj', 'a_p_allm', 'a_p_min', 'p_h_conj', 'p_h_maj', 'p_h_allm', 'p_h_min', 'h_p_conj', 'h_p_maj', 'h_p_allm', 'h_p_min', 'p_p_conj', 'p_p_maj', 'p_p_allm', 'e', 'all_conj', 'all_maj', 'a_p_p_a_conj', 'all_m']
            elif technique == TechniqueType.SECONDARY_DIRECT:
                selected_categories = ['time', 'p_a_conj', 'p_a_maj', 'p_a_allm', 'p_a_min', 'a_p_conj', 'a_p_maj', 'a_p_allm', 'a_p_min', 'p_h_conj', 'p_h_maj', 'p_h_allm', 'p_h_min', 'h_p_conj', 'h_p_maj', 'h_p_allm', 'h_p_min', 'p_p_conj', 'p_p_maj', 'p_p_allm', 'mon_p_conj', 'mon_p_maj', 'mon_p_allm', 'mon_p_min', 'e', 'all_maj', 'all_conj', 'all_m']
            elif technique == TechniqueType.PSSR:
                selected_categories = ['time', 'p_a_conj', 'p_a_maj', 'p_a_allm', 'p_a_min', 'p_h_conj', 'p_h_maj', 'p_h_allm', 'p_h_min', 'p_p_conj', 'p_p_maj', 'p_p_allm', 'mon_p_conj', 'mon_p_maj', 'mon_p_allm', 'mon_p_min', 'e', 'all_conj', 'all_maj', 'p_a_mon_conj']
            elif technique == TechniqueType.TRANSIT:
                selected_categories = ['time', 'p_a_allm', 'p_a_min', 'e', 'p_a_maj', 'p_a_conj']

            filtered_counts = {key: count[key] for key in selected_categories if key in count}
            f.write(f"Row {index + 1}: {filtered_counts}\n")

def count_aspect_groups_txt(filename, flag_count_moon):
    results = []

    with open(f"{filename}.txt", 'r') as infile:
        for line in infile:
            parts = eval(line.strip())
            time = parts[0]
            aspects = parts[1:-1]
            count = parts[-1]

            opp_conj_count = 0
            sqr_tri_sext_count = 0
            minor_count = 0
            moon_conj_opp_count = 0
            moon_sqr_tri_sext_count = 0
            empty_event_count = 0

            for all_aspect in aspects:
                if all_aspect:
                    if all_aspect[0] != '(':
                        empty_event_count += 1

                    if '\n' in all_aspect:
                        individual_aspects_list = all_aspect.split('\n')
                    else:
                        individual_aspects_list = [all_aspect]

                    for asp in individual_aspects_list:
                        if ('sesquisquare' in asp) or ('semisquare' in asp) or ('semisextile' in asp) or ('quincunx' in asp):
                            minor_count += 1
                        elif ('opposition' in asp) or ('conjunction' in asp):
                            if (flag_count_moon and (')) (Moon' in asp)):
                                moon_conj_opp_count += 1
                            opp_conj_count += 1
                        elif ('square' in asp) or ('sextile' in asp) or ('trine' in asp):
                            if (flag_count_moon and (')) (Moon' in asp)):
                                moon_sqr_tri_sext_count += 1
                            sqr_tri_sext_count += 1
                        
            if flag_count_moon:
                results.append([f"{time}, {count}, opp-conj: {opp_conj_count}, sqr-tri-sext: {sqr_tri_sext_count}, major: {sqr_tri_sext_count+opp_conj_count}, minor: {minor_count}, moon-opp-conj: {moon_conj_opp_count}, moon-sqr-tri-sext: {moon_sqr_tri_sext_count}, empty: {empty_event_count}"])                
            else:
                results.append([f"{time}, {count}, opp-conj: {opp_conj_count}, sqr-tri-sext: {sqr_tri_sext_count}, major: {sqr_tri_sext_count+opp_conj_count}, minor: {minor_count}, empty: {empty_event_count}"])                
    print(results)

    with open(f"{filename}COUNT.txt", 'w') as outfile:
        for result in results:
            outfile.write(str(result) + '\n')

def generate_grid_times_manual(filename, list_times, list_dt_events, geo_positions: list[3], type: pd_automate.AspectType, technique: TechniqueType):
    global grid_aspects, date_technique, aspect_type
    date_technique = technique
    aspect_type = type
    temp_list_event = ['Time']

    for i in range(0,len(list_dt_events)):
        temp_list_event.append(f"{i}: {list_dt_events[i][0].strftime('%Y-%m-%d')}")
    temp_list_event.append('Count')
    grid_aspects.append(temp_list_event)
    
    for current_time in list_times:
        print(f"working on: {current_time}....")
        append_grid_acceptable_angles(list_dt_events, julian.to_jd(current_time),geo_positions)
    
    directory = os.path.dirname(filename)
    os.makedirs(directory, exist_ok=True)

    with open(f"{filename}.txt", "w") as file:
        for time in grid_aspects:
            file.write(f"{str(time)}\n")

from datetime import timedelta

def add_timezone_to_24_hours(timezone):
    base_time = timedelta(hours=24)
    adjusted_time = base_time + timedelta(hours=timezone)
    result_hours = adjusted_time.seconds // 3600 % 24
    
    return result_hours

def process_manual_rect_csv(file_path, dt_bday, count_times_wanted, geopos):
    """end date is day of birth 
    timezone is whatever you need to add to UT to get the local time"""
    i_timezone = get_timezone_from_pos(geopos)
    df = pd.read_csv(file_path)
    times = df['Time'][:count_times_wanted].tolist()
    
    datetime_list = get_list_datetime_from_times(times, dt_bday, i_timezone)
    
    return datetime_list

def get_list_datetime_from_times(str_times, dt_bday, i_timezone):
    datetime_list = []

    for time_str in str_times:
        time_obj = datetime.datetime.strptime(time_str, '%H:%M:%S')
        
        if time_obj.hour < add_timezone_to_24_hours(i_timezone):
            if i_timezone <= 0:
                datetime_obj = datetime.datetime.combine(dt_bday, time_obj.time())
            else: 
                datetime_obj = datetime.datetime.combine(dt_bday + timedelta(days=1), time_obj.time())
        else:
            if i_timezone <= 0:
                datetime_obj = datetime.datetime.combine(dt_bday - timedelta(days=1), time_obj.time())
            else:
                datetime_obj = datetime.datetime.combine(dt_bday, time_obj.time())
    
        datetime_list.append(datetime_obj)
    
    return datetime_list

def sort_polaris_times(file_name, file_write＿name, count_times_wanted, threshold=None):
    '''
    created new file with polaris times sorted according to A value the if you give threshold
    removes the lines that do not have enough A points
    '''
    line_count = 0
    with open(file_name, 'r') as file:
        for line in file:
            line_count += 1
    if count_times_wanted > (line_count/2):
        print('Not enough times for that count...')
        count_times_wanted = 50 #assume polaris will always give over 50 peaks?
        return 

    with open(file_name, 'r') as file:
        lines = file.readlines()

    valid_lines = [line for line in lines if len(line.split()) > 3]
    non_valid_lines = [line for line in lines if len(line.split()) <= 3]
    year = non_valid_lines[0].strip()

    sorted_lines = sorted(valid＿lines, key=lambda x: int(x.split()[4]), reverse=True)
    #sorted_lines = sorted(valid_lines, key=lambda x: int(x.split()[4]) + int(x.split()[5]), reverse=True)

    filtered_lines = [
        line.strip()
        for line in sorted_lines
        if len(line.split()) > 5 and int(line.split()[4]) >= threshold
    ]
    filtered_lines = ["\t" + line.strip() + f"\n\t{year}" for line in filtered_lines]

    with open(file_write_name, 'w')  as file:
        for line in filtered_lines:
            file.write(line + '\n')

    return

def process_polaris_times(file_name, count_times_wanted):
    """Processes a text file with date and time entries.
    
    Args:
        file_path (str): Path to the text file.
        count_times_wanted (int): Number of times to process.
        i_timezone (int): Timezone offset to apply.

    Returns:
        List[datetime]: A list of datetime objects.
    """
    line_count = 0
    with open(file_name, 'r') as file:
        for line in file:
            line_count += 1
    if count_times_wanted > (line_count/2):
        print('Not enough times for that count...')
        return 

    with open(file_name, 'r') as file:
        lines = file.readlines()

    datetime_list = []

    for i in range(0, count_times_wanted*2, 2):
        date_str = lines[i].strip()
        date_str = re.sub(r'\s+', ' ', date_str).strip()
        date_str_list = date_str.split(' ')
        date_time_str = lines[i + 1].strip() + ' ' + date_str_list[0] + ' ' + date_str_list[1] + ' ' + date_str_list[2]

        dt = datetime.datetime.strptime(date_time_str, '%Y %d %b %H:%M:%S')
        datetime_list.append(dt)

    return datetime_list

def count_pssr_moon_from_times_events(filename_write, list_datetimes: list, events_list :list, geopos):
    count_array = [('DateTime','count')]
    aspect_type = pd_automate.AspectType.MOON_PRIMARY

    for dtime in list_datetimes:
        count_moon_conj_opp = 0
        for dt_event, event_id, _ in events_list:
            pssr_obj = pssr_auto.PSSR_Auto(dtime, dt_event, None, geopos)
            str_rad_dir_aspects, str_rad_conv_aspects = pssr_obj.get_str_aspects()
            str_all_directed_aspects = str_rad_dir_aspects + str_rad_conv_aspects 
            _, str_acceptable_aspects = pd_automate.count_event_acceptable_aspects(event_id, str_all_directed_aspects, 0, aspect_type)
            list_aspects = str_acceptable_aspects.split('\n')
            temp_arr = []
            for str_aspect in list_aspects:
                try:
                    if pd_automate.is_aspect_conj_opp(str_aspect):
                        temp_arr.append(str_aspect)
                except:
                    pass
            list_aspects = temp_arr
            count_moon_conj_opp += len(list_aspects)
        count_array.append((dtime,count_moon_conj_opp))

    count_array = [count_array[0]] + sorted(count_array[1:], key=lambda x: x[1], reverse=True)

    with open(filename_write, mode='w', newline='') as file:
        writer_csv = csv.writer(file)
        for item in count_array:
            writer_csv.writerow([item[0].strftime('%Y-%m-%d %H:%M:%S') if isinstance(item[0], datetime.datetime) else item[0], item[1]])

def process_datetime_count_csv(filename_read):
    datetime_list = []

    with open(filename_read, mode='r', newline='') as file:
        reader = csv.reader(file)
        next(reader)
        
        for row in reader:
            dt = datetime.datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S')
            datetime_list.append(row[0])

    return datetime_list

def process_time_count_csv(filename_read, dt_bday, geopos):
    time_list = []

    i_timezone = get_timezone_from_pos(geopos)

    with open(filename_read, mode='r', newline='') as file:
        reader = csv.reader(file)
        next(reader)
        
        for row in reader:
            time_list.append(row[0])

    datetime_list = get_list_datetime_from_times(time_list, dt_bday, i_timezone)

    return datetime_list

def get_timezone_from_pos(geopos):
    tf = TimezoneFinder()
    geo_lat = geopos[0]
    geo_long = geopos[1]
    timezone_name = tf.timezone_at(lat=geo_lat, lng=geo_long)
    timezone = pytz.timezone(timezone_name)
    now = datetime.datetime.now(timezone)
    utc_offset = now.utcoffset().total_seconds() / 3600

    return utc_offset

def generate_hourly_datetimes(geopos, input_datetime):
    utc_offset = get_timezone_from_pos(geopos)
    
    start_datetime = input_datetime.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(hours=utc_offset)
    end_datetime = start_datetime + timedelta(days=1)

    hourly_datetimes = []
    current_datetime = start_datetime
    while current_datetime <= end_datetime:
        hourly_datetimes.append(current_datetime)
        current_datetime += timedelta(minutes=5)

    return hourly_datetimes

def delete_rows_below_threshold_counttxt(threshold, file_name): 
    with open(file_name, 'r') as file:
        rows = file.readlines()

    filtered_rows = []
    for row in rows:
        start_index = row.find('{')
        end_index = row.rfind('}')
        if start_index != -1 and end_index != -1:
            row_dict = eval(row[start_index:end_index + 1])
            last_value = list(row_dict.values())[-1]
            
            if isinstance(last_value, (int, float)) and last_value >= threshold:
                filtered_rows.append((last_value, row))

    filtered_rows.sort(reverse=True, key=lambda x: x[0])
    sorted_rows = [row[1] for row in filtered_rows]

    with open(file_name, 'w') as file:
        file.writelines(sorted_rows)
        
#the following is to do with the summing of two files with the same times prim sec rectification verification stuff
# Function to read and process the file
def read_file(file_path):
    data = {}
    with open(file_path, 'r') as f:
        for line in f:
            # Extracting the Row number and the dictionary
            row_number, row_data = line.split(":", 1)
            row_number = row_number.strip()
            row_data = ast.literal_eval(row_data.strip())  # Safely convert the string to a dictionary
            time = row_data['time']
            # Store the all_m value along with the time in the dictionary
            data[time] = row_data['all_m']
    return data

# Function to sum the 'all_m' values from two dictionaries
def sum_all_m(data1, data2):
    result = {}
    # Merge data1 and data2 by summing values for the same time
    for time in data1:
        if time in data2:
            result[time] = data1[time] + data2[time]
    return result

# Function to write the result into a new file
def write_result(file_path, result):
    directory = os.path.dirname(file_path)
    os.makedirs(directory, exist_ok=True)
    
    with open(file_path, 'w') as f:
        for i, (time, summed_value) in enumerate(result.items(), start=1):
            f.write(f"Row {i}: {{'time': '{time}', 'sum_all_m': {summed_value}}}\n")

def sum_sec_prim(prim_filename, sec_filename):
    result_file_path = f"{prim_filename[:-18]}summed_prim_sec.txt"
    prim_data = read_file(prim_filename)
    sec_data = read_file(sec_filename)
    result = sum_all_m(prim_data, sec_data)

    write_result(result_file_path, result)
    
def sanitize_sheet_name(name):
    """Sanitizes a string to be a valid Excel sheet name."""
    # Max length is 31. Invalid chars: []:*?/\\
    name = name.replace(':', '-').replace('/', '_').replace('\\', '_')
    name = name.replace('[', '(').replace(']', ')')
    name = name.replace('*', '_').replace('?', '_')
    return name[:31] # Truncate to 31 chars

def abbreviate_aspect_string(aspect_line):
    """
    Transforms an aspect string to an abbreviated format including direction.
    Original: (Planet1,Deg,(Dir1)) (Planet2,Deg,(Dir2)) (AspectName,Orb')
    Target:   ABB1 (Dir1) DEGREE ABB2 (Dir2) (Orb')
    """
    # Regex to capture the main parts including directions
    # Group 1: Planet1 Name
    # Group 2: Planet1 Direction (e.g., 'c', 'r', 'd')
    # Group 3: Planet2 Name
    # Group 4: Planet2 Direction (e.g., 'c', 'r', 'd')
    # Group 5: Aspect Name
    # Group 6: Orb (e.g., "0.06'")
    pattern = re.compile(
        r"\(([^,]+),[0-9.]+,\(([^)]+)\)\)"  # Planet 1 block (captures name, then direction)
        r"\s+"
        r"\(([^,]+),[0-9.]+,\(([^)]+)\)\)"  # Planet 2 block (captures name, then direction)
        r"\s+"
        r"\(([^,]+),([0-9.]+'\))"          # Aspect block (captures aspect name and orb)
    )
    match = pattern.match(aspect_line.strip())
    
    if match:
        p1_name_full = match.group(1)
        p1_dir = match.group(2) # Captured direction for P1 is it converse or radix etc.
        p2_name_full = match.group(3)
        p2_dir = match.group(4) # Captured direction for P2
        aspect_name_full = match.group(5).lower()
        orb_str = match.group(6)[:-1] # This is "0.06'" very ugly fix for why there where 2 brackets in the display string so just forced out the bracket from the match group

        p1_abbr = PLANET_ABBREVIATIONS.get(p1_name_full, p1_name_full[:3].upper()) # Default to first 3 chars if not found
        p2_abbr = PLANET_ABBREVIATIONS.get(p2_name_full, p2_name_full[:3].upper())
        aspect_deg = ALL_ASPECTS.get(aspect_name_full)[0]

        if aspect_deg is not None:
            return f"{p1_abbr} ({p1_dir}) {aspect_deg} {p2_abbr} ({p2_dir}) ({orb_str})"
        else:
            logging.warning(f"Aspect name '{aspect_name_full}' not found in ASPECT_DEGREES for line: {aspect_line}")
            return aspect_line # Return original if aspect name not found
    else:
        # If the line doesn't match the expected 3-part structure, return it as is
        # This handles Natal listings, headers, or malformed lines gracefully.
        logging.debug(f"Line did not match abbreviation pattern: {aspect_line}")
        return aspect_line

def create_analysis_workbook():
    # 1. Get User Input for JSON filename
    while True:
        json_filename_input = input(f"Enter the JSON data filename (e.g., charlie chaplin.json, found in '{DATA_INPUT_DIR}'): ").strip()
        json_filepath = os.path.join(DATA_INPUT_DIR, json_filename_input)
        if os.path.exists(json_filepath):
            break
        else:
            logging.error(f"File not found: {json_filepath}. Please try again.")

    # 2. Derive base name for .txt files and Excel output
    base_name_for_txt = os.path.splitext(json_filename_input)[0].replace(' ', '_')
    excel_output_filename = f"{base_name_for_txt}_rectification_analysis.xlsx"
    logging.info(f"Derived base name for text files: {base_name_for_txt}")
    logging.info(f"Output Excel file will be: {excel_output_filename}")

    # 3. Technique Selection
    available_techniques = aTechniqueType.get_all_techniques()
    print("\nAvailable Techniques:")
    for index, name in sorted(available_techniques.items()):
        print(f"{index}: {name}")

    while True:
        try:
            selected_indices_str = input("Enter the numbers of the techniques to process, separated by commas (e.g., 0,1,6): ").strip()
            selected_indices = [int(s.strip()) for s in selected_indices_str.split(',') if s.strip()]
            if not selected_indices:
                raise ValueError("No techniques selected.")
            # Validate indices
            valid_selection = True
            for idx in selected_indices:
                if idx not in available_techniques:
                    logging.error(f"Invalid technique index: {idx}")
                    valid_selection = False
                    break
            if valid_selection:
                break
        except ValueError as e:
            logging.error(f"Invalid input. Please enter numbers separated by commas. Error: {e}")

    selected_technique_names = [get_technique_name(idx) for idx in selected_indices]
    logging.info(f"Selected techniques for processing: {', '.join(selected_technique_names)}")

    # 4. Load or Create Workbook
    workbook = None
    if os.path.exists(excel_output_filename):
        logging.info(f"Loading existing workbook: {excel_output_filename}")
        try:
            workbook = openpyxl.load_workbook(excel_output_filename)
        except Exception as e:
            logging.error(f"Could not load existing workbook {excel_output_filename}: {e}. A new one will be created.")
            workbook = openpyxl.Workbook() # Create new if loading fails
    else:
        logging.info(f"Creating new workbook: {excel_output_filename}")
        workbook = openpyxl.Workbook()
        # Remove default sheet if a new workbook is created
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
            workbook.remove(workbook["Sheet"])


    # 5. Parse JSON for Event Headers
    event_headers_formatted = [] # For Excel display: "EVENT_TYPE YYYY-MM-DD"
    event_data_from_json = [] # To store {formatted_header: original_full_event_string} mapping

    try:
        with open(json_filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            json_events = data.get("list_of_events", [])
            temp_event_list_for_sorting = [] # For sorting events by datetime
            for event_detail in json_events:
                dt_str = event_detail.get("datetime", "")
                event_type = event_detail.get("event_type", "UNKNOWN_EVENT")
                event_geopos = event_detail.get("geopos", []) # Not used in header but part of original_full_event_string
                # Find the event index from the original source, if possible, or generate one
                # This part depends on how event_id is stored/used in your .txt files
                # For simplicity, let's assume the event type from JSON is sufficient with date.
                # The original event string in .txt files is like "YYYY-MM-DDTHH:MM:SS, TYPE, ID, [GEO]"
                # We need to match based on TYPE and YYYY-MM-DD part of the datetime from JSON.

                # Create the formatted header
                try:
                    event_dt_obj = datetime.datetime.fromisoformat(dt_str)
                    formatted_header = f"{event_type} {event_dt_obj.strftime('%Y-%m-%d')}"
                    temp_event_list_for_sorting.append((event_dt_obj, formatted_header))

                    # Store mapping for later lookup. The key in all_data will be the *exact* string
                    # from the "Event: ..." line in the .txt file.
                    # We need to be able to find this original string based on our formatted_header.
                    # This is the trickiest part. For now, let's assume we will iterate through all_data's event keys
                    # and try to match them to the formatted_header.

                except ValueError:
                    logging.warning(f"Could not parse datetime for event: {event_detail}")
                    # Add with max date to sort invalid ones last
                    formatted_header = f"{event_type} INVALID_DATE"
                    temp_event_list_for_sorting.append((datetime.max, formatted_header))

            temp_event_list_for_sorting.sort()

            for _, formatted_header in temp_event_list_for_sorting:
                event_headers_formatted.append(formatted_header)
            #event headers is the sorted list of temp event list but with the date after string label    
                
            if not event_headers_formatted:
                logging.error("No events found in JSON file. Cannot proceed.")
                return

    except Exception as e:
        logging.error(f"Error reading or parsing JSON file {json_filepath}: {e}")
        return

    # 6. Find and Parse all relevant .txt files
    all_parsed_data = {} # { "filename_datetime": { "original_event_string": { "TechniqueName": [aspects] } } }
    filename_datetimes = []

    for txt_filename in os.listdir(SELECTIONS_DIR):
        if txt_filename.startswith(base_name_for_txt + "_") and txt_filename.endswith(".txt"):
            full_txt_filepath = os.path.join(SELECTIONS_DIR, txt_filename)
            # Extract datetime part from filename: base_name_YYYY-MM-DD_HH-MM-SS.txt
            try:
                datetime_part_from_filename = txt_filename[len(base_name_for_txt) + 1:-4] # Remove prefix and .txt
                # Further sanitize/validate datetime_part_from_filename if needed
                datetime.datetime.strptime(datetime_part_from_filename, "%Y-%m-%d_%H-%M-%S") # Validate format
                filename_datetimes.append(datetime_part_from_filename)
            except ValueError:
                logging.warning(f"Could not parse datetime from filename {txt_filename}. Skipping.")
                continue

            parsed_selections = parse_selection_file(full_txt_filepath)
            if parsed_selections:
                all_parsed_data[datetime_part_from_filename] = parsed_selections
            else:
                logging.warning(f"No data parsed from {full_txt_filepath}")

    if not all_parsed_data:
        logging.error(f"No .txt selection files found or parsed for base name '{base_name_for_txt}' in '{SELECTIONS_DIR}'.")
        return

    filename_datetimes = sorted(list(set(filename_datetimes))) # Unique datetimes

    # 7. Populate Excel Sheets
    current_timestamp_for_sheets = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    for technique_name_to_process in selected_technique_names: # Iterate only selected techniques
        new_sheet_name_base = technique_name_to_process
        new_sheet_name_dated = sanitize_sheet_name(f"{new_sheet_name_base}_{current_timestamp_for_sheets}")

        # Remove existing sheet with the same dated name if it somehow exists (should be rare)
        if new_sheet_name_dated in workbook.sheetnames:
            workbook.remove(workbook[new_sheet_name_dated])

        sheet = workbook.create_sheet(title=new_sheet_name_dated)
        logging.info(f"Processing data for sheet: {new_sheet_name_dated}")

        # Write Headers
        sheet['A1'] = "Datetime \\ Event" # Label for A1
        sheet['A1'].font = Font(bold=True)
        for col_idx, header_text in enumerate(event_headers_formatted, start=2): # Start at col B
            col_letter = get_column_letter(col_idx)
            sheet[f"{col_letter}1"] = header_text
            sheet[f"{col_letter}1"].font = Font(bold=True)
            sheet[f"{col_letter}1"].alignment = Alignment(wrap_text=True) # Wrap header text
            sheet.column_dimensions[col_letter].width = 24 # Set column width


        # Write Datetime Row Labels and Data
        for row_idx, filename_dt in enumerate(filename_datetimes, start=2): # Start at row 2
            sheet[f"A{row_idx}"] = filename_dt
            sheet[f"A{row_idx}"].font = Font(bold=True)

            # Get data for this specific filename_datetime
            data_for_this_dt = all_parsed_data.get(filename_dt, {})

            for col_idx, formatted_event_header in enumerate(event_headers_formatted, start=2):
                # Find the original_full_event_string that matches this formatted_event_header
                # This requires matching the "EVENT_TYPE YYYY-MM-DD" part
                target_event_type = formatted_event_header.split(' ')[0]
                target_event_date = formatted_event_header.split(' ')[1]

                found_original_event_key = None
                for original_event_key in data_for_this_dt.keys():
                    # Parse the original_event_key (e.g., "2004-12-01T12:00:00, MOVE_HOME, ...")
                    try:
                        orig_dt_str = original_event_key.split(',')[0].strip()
                        orig_type_str = original_event_key.split(',')[1].strip()
                        orig_dt_obj = datetime.datetime.fromisoformat(orig_dt_str)

                        if orig_type_str == target_event_type and orig_dt_obj.strftime('%Y-%m-%d') == target_event_date:
                            found_original_event_key = original_event_key
                            break
                    except Exception: # Broad exception for parsing robustness
                        continue # Skip if original_event_key is malformed

                if found_original_event_key:
                    aspect_data_for_cell = data_for_this_dt.get(found_original_event_key, {}).get(technique_name_to_process, [])
                    if aspect_data_for_cell:
                        # Abbreviate each aspect string in the list
                        abbreviated_aspects = [abbreviate_aspect_string(aspect) for aspect in aspect_data_for_cell]
                        sheet.cell(row=row_idx, column=col_idx).value = "\n".join(abbreviated_aspects)
                        sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        sheet.cell(row=row_idx, column=col_idx).value = "" # Empty if no aspects for this combo
                else:
                    sheet.cell(row=row_idx, column=col_idx).value = "" # Empty if event not found in this .txt file's dzata

        sheet.column_dimensions['A'].width = 25 # Width for datetime column

    # 8. Save Workbook
    try:
        workbook.save(excel_output_filename)
        logging.info(f"Successfully saved Excel file: {excel_output_filename}")
    except Exception as e:
        logging.error(f"Error saving Excel file {excel_output_filename}: {e}")

#sort_polaris_times(r'data_times\25_01_07_ingtea rect 5 to 8.txt', 'data_times/25_01_07_ingtea sorted max a 3 rect.txt',58, 3)
#delete_rows_below_threshold_counttxt(0,'data_rect/30_11_24_Ingtea_v1/30_11_24_2000-03-11_pssrCOUNT.txt')
#create_analysis_workbook()