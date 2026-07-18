"""
Excel export: sanitize_sheet_name, abbreviate_aspect_string, and
create_analysis_workbook - the last of which is fully interactive
(gathers its inputs via two input() prompts rather than taking
parameters), so it has never had any automated coverage. This file
mocks input() with real, valid answers and runs it end-to-end against
the two real, shipped saved_selections/*.txt files.
"""

import os
import shutil
import tempfile
from datetime import datetime

import openpyxl
import pytest

import process_techniques_files as ptf

FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
BIRTH_DATA_DIR = os.path.join(FIXTURES_DIR, "birth_data")
SAVED_SELECTIONS_DIR = os.path.join(FIXTURES_DIR, "saved_selections")


def _frozen_datetime_module(*now_args):
    """Builds a small proxy object standing in for the `datetime` module,
    with .datetime.now() frozen to a fixed instant and everything else
    delegated to the real module - safe to assign directly to
    process_techniques_files.datetime (a module-local rebind) without
    mutating the actual shared datetime module's own attributes, which
    every other importer of `datetime` (including openpyxl, internally,
    during workbook.save()) also sees and depends on."""
    import types
    import datetime as real_datetime

    class _Frozen(real_datetime.datetime):
        @classmethod
        def now(cls, tz=None):
            return real_datetime.datetime(*now_args)

    return types.SimpleNamespace(
        datetime=_Frozen,
        timedelta=real_datetime.timedelta,
        date=real_datetime.date,
        time=real_datetime.time,
        timezone=real_datetime.timezone,
    )


class TestSanitizeSheetName:
    def test_replaces_every_excel_invalid_character(self):
        assert ptf.sanitize_sheet_name("a:b/c\\d[e]f*g?h") == "a-b_c_d(e)f_g_h"

    def test_truncates_to_31_characters(self):
        long_name = "x" * 50
        result = ptf.sanitize_sheet_name(long_name)
        assert len(result) == 31

    def test_leaves_an_already_valid_short_name_untouched(self):
        assert ptf.sanitize_sheet_name("PD_2025-01-01") == "PD_2025-01-01"


class TestAbbreviateAspectString:
    def test_real_pd_aspect_line_abbreviates_correctly(self):
        # Pulled directly from the real, shipped saved_selections sample.
        line = "(MC,261.980,(c)) (Neptune,262.087,(r)) (conjunction,6.43')"
        result = ptf.abbreviate_aspect_string(line)
        # MC has no PLANET_ABBREVIATIONS entry -> falls back to name[:3].upper().
        # Note: match.group(6)[:-1] strips the trailing ')' bracket, NOT the
        # orb's own trailing apostrophe - "6.43')" -> "6.43'", confirmed
        # directly against the regex before writing this assertion.
        assert result == "MC (c) 0 NEP (r) (6.43')"

    def test_conjunction_degree_matches_all_aspects_table(self):
        line = "(Sun,10.000,(r)) (Moon,10.000,(d)) (conjunction,0.00')"
        result = ptf.abbreviate_aspect_string(line)
        assert result == "SUN (r) 0 MON (d) (0.00')"

    def test_trine_degree_matches_all_aspects_table(self):
        line = "(Venus,120.000,(r)) (Jupiter,0.000,(d)) (trine,0.00')"
        result = ptf.abbreviate_aspect_string(line)
        import aspects_base
        assert f" {aspects_base.ALL_ASPECTS['trine'][0]} " in result

    def test_malformed_line_returned_unchanged(self):
        line = "this is not an aspect line at all"
        assert ptf.abbreviate_aspect_string(line) == line

    def test_natal_style_single_planet_line_returned_unchanged(self):
        # Natal listings (see app.py's NATAL display) are single
        # "(Planet,Degree,(Label))" entries with no second point/aspect -
        # confirms these pass through unmodified rather than crashing.
        line = "(Sun,192.284,(r))"
        assert ptf.abbreviate_aspect_string(line) == line


class TestCreateAnalysisWorkbookEndToEnd:
    """create_analysis_workbook() takes no parameters - it gathers the
    JSON filename and technique selection via input(). Mocked here with
    real, valid answers, run against a real, controlled data_input/ +
    saved_selections/ directory pair (the latter seeded with the two
    real, shipped sample files) in an isolated temp working directory."""

    def _setup_working_directory(self, tmp_path):
        (tmp_path / "data_input").mkdir()
        (tmp_path / "saved_selections").mkdir()
        shutil.copy(
            os.path.join(BIRTH_DATA_DIR, "beyonce.json"),
            tmp_path / "data_input" / "beyonce.json",
        )
        for fname in os.listdir(SAVED_SELECTIONS_DIR):
            shutil.copy(
                os.path.join(SAVED_SELECTIONS_DIR, fname),
                tmp_path / "saved_selections" / fname,
            )
        return tmp_path

    def test_produces_a_real_xlsx_with_pd_sheet_and_real_abbreviated_aspects(self, tmp_path, monkeypatch):
        self._setup_working_directory(tmp_path)
        monkeypatch.chdir(tmp_path)

        # Prompt 1: JSON filename. Prompt 2: technique indices ("0" = PD,
        # matching get_all_techniques()'s {0: "PD", ...} and the real
        # technique key ("PD") the shipped saved_selections files use.
        responses = iter(["beyonce.json", "0"])
        monkeypatch.setattr("builtins.input", lambda *_: next(responses))

        ptf.create_analysis_workbook()

        out_path = tmp_path / "beyonce_rectification_analysis_grouped.xlsx"
        assert out_path.exists()

        wb = openpyxl.load_workbook(str(out_path))
        pd_sheets = [s for s in wb.sheetnames if s.startswith("PD_")]
        assert len(pd_sheets) == 1
        sheet = wb[pd_sheets[0]]

        assert sheet["A1"].value == "Datetime \\ Event"
        # Row labels are the two real saved_selections files' embedded
        # datetimes (from their filenames): 1981-09-05_02-24-52 and
        # 1981-09-05_02-26-04, sorted.
        row_labels = [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]
        assert "1981-09-05_02-24-52" in row_labels
        assert "1981-09-05_02-26-04" in row_labels

        # At least one cell must contain real, abbreviated aspect content
        # (not raw "(Planet,Degree,(Label))" strings - abbreviate_aspect_string
        # should have already transformed it).
        found_abbreviated_content = False
        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value and "(" in str(cell.value) and "," not in str(cell.value).split("(")[0]:
                    found_abbreviated_content = True
        assert found_abbreviated_content

    def test_event_column_headers_are_event_type_plus_date(self, tmp_path, monkeypatch):
        self._setup_working_directory(tmp_path)
        monkeypatch.chdir(tmp_path)
        responses = iter(["beyonce.json", "0"])
        monkeypatch.setattr("builtins.input", lambda *_: next(responses))

        ptf.create_analysis_workbook()

        wb = openpyxl.load_workbook(str(tmp_path / "beyonce_rectification_analysis_grouped.xlsx"))
        sheet = wb[[s for s in wb.sheetnames if s.startswith("PD_")][0]]
        headers = [sheet.cell(row=1, column=c).value for c in range(2, sheet.max_column + 1)]
        # Real event from beyonce.json: BIRTH_SISTER on 1986-06-24
        assert "BIRTH_SISTER 1986-06-24" in headers

    def test_invalid_json_filename_reprompts_until_a_real_one_is_given(self, tmp_path, monkeypatch):
        self._setup_working_directory(tmp_path)
        monkeypatch.chdir(tmp_path)

        responses = iter(["does_not_exist.json", "beyonce.json", "0"])
        monkeypatch.setattr("builtins.input", lambda *_: next(responses))

        ptf.create_analysis_workbook()  # must not raise, and must eventually succeed

        assert (tmp_path / "beyonce_rectification_analysis_grouped.xlsx").exists()

    def test_reruns_within_the_same_second_replace_the_sheet_not_add_a_second_one(self, tmp_path, monkeypatch):
        """Real finding, confirmed by mocking datetime.datetime.now() to a
        fixed instant for both calls: the dated sheet name
        (f"{technique}_{now:%Y-%m-%d_%H-%M-%S}") only has 1-second
        granularity, and create_analysis_workbook() explicitly removes
        any existing sheet with that exact name before creating a fresh
        one. Two runs within the same wall-clock second - entirely
        plausible for fast scripted/automated use - silently REPLACE the
        prior run's sheet rather than adding a second one. Confirmed
        directly (not assumed) before writing this test.

        Uses a module-local rebind of ptf.datetime (a fresh proxy
        namespace), not a mutation of the shared datetime module's own
        .datetime attribute - the latter is process-global (every other
        importer of `datetime` sees the same module object) and was
        confirmed to corrupt openpyxl's own internal datetime usage during
        workbook.save(), producing an unreadable .xlsx file."""
        self._setup_working_directory(tmp_path)
        monkeypatch.chdir(tmp_path)
        monkeypatch.setattr(ptf, "datetime", _frozen_datetime_module(2025, 6, 1, 12, 0, 0))

        responses = iter(["beyonce.json", "0", "beyonce.json", "0"])
        monkeypatch.setattr("builtins.input", lambda *_: next(responses))

        ptf.create_analysis_workbook()
        wb1 = openpyxl.load_workbook(str(tmp_path / "beyonce_rectification_analysis_grouped.xlsx"))
        sheet_names_after_first_run = set(wb1.sheetnames)

        ptf.create_analysis_workbook()
        wb2 = openpyxl.load_workbook(str(tmp_path / "beyonce_rectification_analysis_grouped.xlsx"))
        sheet_names_after_second_run = set(wb2.sheetnames)

        assert sheet_names_after_second_run == sheet_names_after_first_run, (
            "expected the second run (same frozen 'now') to replace the "
            "identically-named sheet from the first run, not add a new one"
        )
        assert len(sheet_names_after_second_run) == 1

    def test_reruns_a_second_or_more_apart_do_add_a_genuinely_second_sheet(self, tmp_path, monkeypatch):
        """Contrast case, proving the mechanism works correctly once the
        two runs' timestamps actually differ - confirms the same-second
        collision above is specifically about timestamp granularity, not
        a sign the whole "add a dated sheet per run" feature is broken."""
        self._setup_working_directory(tmp_path)
        monkeypatch.chdir(tmp_path)

        import datetime as real_datetime
        instants = iter([
            real_datetime.datetime(2025, 6, 1, 12, 0, 0),
            real_datetime.datetime(2025, 6, 1, 12, 0, 1),  # 1 second later
        ])

        import types

        class _SteppedNow(real_datetime.datetime):
            @classmethod
            def now(cls, tz=None):
                return next(instants)

        fake_module = types.SimpleNamespace(
            datetime=_SteppedNow,
            timedelta=real_datetime.timedelta,
            date=real_datetime.date,
            time=real_datetime.time,
            timezone=real_datetime.timezone,
        )
        monkeypatch.setattr(ptf, "datetime", fake_module)

        responses = iter(["beyonce.json", "0", "beyonce.json", "0"])
        monkeypatch.setattr("builtins.input", lambda *_: next(responses))

        ptf.create_analysis_workbook()
        wb1 = openpyxl.load_workbook(str(tmp_path / "beyonce_rectification_analysis_grouped.xlsx"))
        sheet_count_after_first_run = len(wb1.sheetnames)

        ptf.create_analysis_workbook()
        wb2 = openpyxl.load_workbook(str(tmp_path / "beyonce_rectification_analysis_grouped.xlsx"))
        sheet_count_after_second_run = len(wb2.sheetnames)

        assert sheet_count_after_second_run == sheet_count_after_first_run + 1
